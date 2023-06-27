import tkinter as tk
from tkinter import ttk
from datetime import datetime, timedelta
from openpyxl import Workbook, load_workbook
from tkinter import messagebox

def save_data():
    name = name_entry.get()
    birthday = birthday_entry.get()

    wb = load_workbook('data.xlsx')
    ws = wb.active
    row = ws.max_row + 1
    ws.cell(row=row, column=1).value = name
    ws.cell(row=row, column=2).value = birthday
    wb.save('data.xlsx')

    refresh_names_combo()  # 保存数据后立即刷新选择框

def calculate_age():
    selected_name = names_combo.get()

    wb = load_workbook('data.xlsx')
    ws = wb.active
    for row in range(1, ws.max_row + 1):
        if ws.cell(row=row, column=1).value == selected_name:
            birthday_str = ws.cell(row=row, column=2).value
            birthday = datetime.strptime(birthday_str, '%Y-%m-%d').date()
            second_date_str = second_date_entry.get()
            second_date = datetime.strptime(second_date_str, '%Y-%m-%d').date()
            age_delta = second_date - birthday
            years = age_delta.days // 365
            months = (age_delta.days % 365) // 30
            result = f"{years}年 零 {months}个月"
            result_label.config(text=result)
            ws.cell(row=row, column=3).value = second_date_str
            ws.cell(row=row, column=4).value = result
            break

    wb.save('data.xlsx')

def refresh_names_combo():
    names_list.clear()  # 清空选择框内容

    wb = load_workbook('data.xlsx')
    ws = wb.active
    for row in range(1, ws.max_row + 1):
        name = ws.cell(row=row, column=1).value
        if name is not None:  # 移除"None"选项
            names_list.append(name)

    names_combo['values'] = names_list  # 更新选择框的值

def clear_data():
    selected_name = names_combo.get()

    if selected_name == "":
        return

    confirmation = messagebox.askquestion("确认", "是否确认清除？")
    if confirmation == "yes":
        wb = load_workbook('data.xlsx')
        ws = wb.active
        for row in range(1, ws.max_row + 1):
            if ws.cell(row=row, column=1).value == selected_name:
                ws.delete_rows(row)
                wb.save('data.xlsx')
                refresh_names_combo()
                name_entry.delete(0, tk.END)
                birthday_entry.delete(0, tk.END)
                second_date_entry.delete(0, tk.END)  # 清空第二个日期输入框的内容
                result_label.config(text="")  # 清空年龄差结果标签的内容
                names_combo.set("")  # 清空选择框的当前选择
                messagebox.showinfo("成功", "已成功清除数据")
                break

window = tk.Tk()
window.title("生日计算器")
window.geometry("400x500")

name_label = tk.Label(window, text="名字：")
name_label.pack()

name_entry = tk.Entry(window)
name_entry.pack()

birthday_label = tk.Label(window, text="生日（YYYY-MM-DD）：")
birthday_label.pack()

birthday_entry = tk.Entry(window)
birthday_entry.pack()

save_button = tk.Button(window, text="保存名字和生日", command=save_data)
save_button.pack(pady=10)

names_label = tk.Label(window, text="选择名字：")
names_label.pack()
names_list = []
wb = load_workbook('data.xlsx')
ws = wb.active
for row in range(1, ws.max_row + 1):
    name = ws.cell(row=row, column=1).value
    if name is not None:
        names_list.append(name)


names_combo = ttk.Combobox(window, values=names_list, state="readonly") # 改成只读，避免光标输入
names_combo.pack()

second_date_label = tk.Label(window, text="第二个日期（YYYY-MM-DD）：")
second_date_label.pack()

second_date_entry = tk.Entry(window)
second_date_entry.pack()

calculate_button = tk.Button(window, text="计算年龄差", command=calculate_age)
calculate_button.pack(pady=10)

result_label = tk.Label(window)
result_label.pack()

clear_button = tk.Button(window, text="清除", command=clear_data)
clear_button.pack(pady=10)

window.mainloop()
